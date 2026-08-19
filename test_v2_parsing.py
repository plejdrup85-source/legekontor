from io import BytesIO

import pandas as pd
import pytest
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from v2.parsing import parse_xlsx_rows


def _structured_workbook() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    main_fill = PatternFill("solid", fgColor="4472C4")
    subsection_fill = PatternFill("solid", fgColor="D9EAF7")

    sheet.append(["Varer   ", None, "GyMo"])
    sheet.append(["UNDERSØKELSESROM", "Varenr", None])
    sheet["A2"].fill = main_fill
    sheet["A2"].font = Font(bold=True)
    sheet.append(["Metall varer", None, None])
    sheet["A3"].fill = subsection_fill
    sheet["A3"].font = Font(bold=True)
    sheet.append(["Spekulum med æøå", 10001, 19.5])
    sheet.append(["Produkt uten varenummer", None, 7.0])
    sheet.append(["Produkt uten pris", 10002, None])
    sheet.append(["Produkt uten varenummer og pris", None, None])
    sheet.append(["SUM", None, "=SUM(C4:C7)"])
    sheet.append(["LAB", None, None])
    sheet["A9"].fill = main_fill
    sheet["A9"].font = Font(bold=True)
    sheet.append(["Større utstyr", None, None])
    sheet["A10"].fill = subsection_fill
    sheet.append(["Måleinstrument", 10003, 100.0])

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def test_xlsx_multiline_headers_and_styled_sections_produce_only_products():
    rows = parse_xlsx_rows(_structured_workbook(), "strukturert.xlsx")

    assert len(rows) == 5
    assert [row["description"] for row in rows] == [
        "Spekulum med æøå",
        "Produkt uten varenummer",
        "Produkt uten pris",
        "Produkt uten varenummer og pris",
        "Måleinstrument",
    ]
    assert [row["competitor_artnr"] for row in rows] == ["10001", "", "10002", "", "10003"]
    assert [row["price_before_discount"] for row in rows] == [19.5, 7.0, None, None, 100.0]
    assert {row["competitor"] for row in rows} == {"GyMo"}
    assert [(row["section"], row["subsection"]) for row in rows] == [
        ("UNDERSØKELSESROM", "Metall varer"),
        ("UNDERSØKELSESROM", "Metall varer"),
        ("UNDERSØKELSESROM", "Metall varer"),
        ("UNDERSØKELSESROM", "Metall varer"),
        ("LAB", "Større utstyr"),
    ]


def test_xlsx_multiline_headers_do_not_guess_between_numeric_columns():
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Varer", None, "Ukjent antall", "Ukjent leverandørpris"])
    sheet.append([None, "Varenr", None, None])
    sheet.append(["Produkt A", 10001, 2, 99.0])
    sheet.append(["Produkt B", 10002, 3, 149.0])
    output = BytesIO()
    workbook.save(output)

    rows = parse_xlsx_rows(output.getvalue(), "tvetydige-kolonner.xlsx")

    assert [row["price_before_discount"] for row in rows] == [None, None]
    assert [row["calculated_unit_price"] for row in rows] == [None, None]
    assert [row["competitor_line_amount"] for row in rows] == [None, None]
    assert {row["competitor"] for row in rows} == {""}


@pytest.mark.parametrize("inventory_header", ["Lagerbeholdning", "Bestand"])
def test_xlsx_multiline_headers_do_not_treat_inventory_as_price(inventory_header):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Varer", None, inventory_header])
    sheet.append([None, "Varenr", None])
    sheet.append(["Produkt A", 10001, 12])
    sheet.append(["Produkt B", 10002, 8])
    output = BytesIO()
    workbook.save(output)

    rows = parse_xlsx_rows(output.getvalue(), "lagerkolonne.xlsx")

    assert [row["price_before_discount"] for row in rows] == [None, None]
    assert [row["calculated_unit_price"] for row in rows] == [None, None]
    assert [row["competitor_line_amount"] for row in rows] == [None, None]
    assert {row["competitor"] for row in rows} == {""}


def test_xlsx_multiline_bold_product_with_unknown_inventory_data_is_preserved():
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Varer", None, "Bestand"])
    sheet.append([None, "Varenr", None])
    sheet.append(["Produkt A", None, 12])
    sheet["A3"].font = Font(bold=True)
    output = BytesIO()
    workbook.save(output)

    rows = parse_xlsx_rows(output.getvalue(), "fet-vare-med-bestand.xlsx")

    assert len(rows) == 1
    assert rows[0]["description"] == "Produkt A"
    assert rows[0]["competitor_artnr"] == ""
    assert rows[0]["price_before_discount"] is None
    assert rows[0]["calculated_unit_price"] is None
    assert rows[0]["competitor_line_amount"] is None
    assert rows[0]["competitor"] == ""


def test_xlsx_multiline_bold_subtotal_with_formula_is_not_a_product():
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Varer", None, "GyMo"])
    sheet.append([None, "Varenr", None])
    sheet.append(["Subtotal", None, "=SUM(C4:C4)"])
    sheet["A3"].font = Font(bold=True)
    sheet.append(["Produkt A", 10001, 99.0])
    output = BytesIO()
    workbook.save(output)

    rows = parse_xlsx_rows(output.getvalue(), "subtotal-formel.xlsx")

    assert [row["description"] for row in rows] == ["Produkt A"]
    assert rows[0]["price_before_discount"] == 99.0
    assert rows[0]["competitor"] == "GyMo"


@pytest.mark.parametrize("total_label", ["Subtotal", "Delsum"])
def test_xlsx_multiline_bold_total_with_constant_inventory_is_not_a_product(
    total_label,
):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Varer", None, "Bestand"])
    sheet.append([None, "Varenr", None])
    sheet.append([total_label, None, 12])
    sheet["A3"].font = Font(bold=True)
    sheet.append(["Produkt A", 10001, 7])
    output = BytesIO()
    workbook.save(output)

    rows = parse_xlsx_rows(output.getvalue(), "konstant-delsum.xlsx")

    assert [row["description"] for row in rows] == ["Produkt A"]
    assert rows[0]["price_before_discount"] is None
    assert rows[0]["competitor"] == ""


@pytest.mark.parametrize("total_label", ["Subtotal", "Delsum"])
@pytest.mark.parametrize("total_value", [12, "=SUM(C5:C5)"])
@pytest.mark.parametrize("is_bold", [False, True])
def test_xlsx_multiline_total_before_late_source_does_not_end_header_block(
    total_label,
    total_value,
    is_bold,
):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Varer", None, None])
    sheet.append([None, "Varenr", None])
    sheet.append([total_label, None, total_value])
    sheet["A3"].font = Font(bold=is_bold)
    sheet.append([None, None, "GyMo"])
    sheet.append(["Produkt A", 10001, 99.0])
    output = BytesIO()
    workbook.save(output)

    rows = parse_xlsx_rows(output.getvalue(), "delsum-for-sen-kilde.xlsx")

    assert [row["description"] for row in rows] == ["Produkt A"]
    assert rows[0]["price_before_discount"] == 99.0
    assert rows[0]["competitor"] == "GyMo"


def test_xlsx_multiline_product_named_subtotal_with_artnr_is_preserved():
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Varer", None, "GyMo"])
    sheet.append([None, "Varenr", None])
    sheet.append(["Subtotal", 10001, 99.0])
    output = BytesIO()
    workbook.save(output)

    rows = parse_xlsx_rows(output.getvalue(), "produkt-med-totalnavn.xlsx")

    assert [row["description"] for row in rows] == ["Subtotal"]
    assert rows[0]["competitor_artnr"] == "10001"
    assert rows[0]["price_before_discount"] == 99.0
    assert rows[0]["competitor"] == "GyMo"


@pytest.mark.parametrize("product_name", ["Subtotal", "Delsum"])
def test_xlsx_multiline_total_named_product_with_artnr_survives_other_formula(
    product_name,
):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Varer", None, "GyMo", None])
    sheet.append([None, "Varenr", None, None])
    sheet.append([product_name, 10001, 99.0, "=SUM(C3:C3)"])
    output = BytesIO()
    workbook.save(output)

    rows = parse_xlsx_rows(output.getvalue(), "totalnavn-med-annen-formel.xlsx")

    assert [row["description"] for row in rows] == [product_name]
    assert rows[0]["competitor_artnr"] == "10001"
    assert rows[0]["price_before_discount"] == 99.0
    assert rows[0]["competitor"] == "GyMo"


def test_xlsx_multiline_total_with_formula_in_artnr_cell_is_filtered():
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Varer", None, "GyMo"])
    sheet.append([None, "Varenr", None])
    sheet.append(["Subtotal", "=COUNT(B4:B4)", 12])
    sheet.append(["Produkt A", 10001, 99.0])
    output = BytesIO()
    workbook.save(output)

    rows = parse_xlsx_rows(output.getvalue(), "total-med-formelvarenummer.xlsx")

    assert [row["description"] for row in rows] == ["Produkt A"]
    assert rows[0]["price_before_discount"] == 99.0
    assert rows[0]["competitor"] == "GyMo"


def test_xlsx_multiline_bold_category_with_text_data_is_not_a_product():
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Varer", None, None])
    sheet.append([None, "Varenr", None])
    sheet.append(["Kategori", None, "Bestand"])
    sheet["A3"].font = Font(bold=True)
    sheet.append(["Produkt A", 10001, 12])
    output = BytesIO()
    workbook.save(output)

    rows = parse_xlsx_rows(output.getvalue(), "kategori-tekstheader.xlsx")

    assert [row["description"] for row in rows] == ["Produkt A"]
    assert rows[0]["price_before_discount"] is None
    assert rows[0]["competitor"] == ""


def test_xlsx_multiline_quantity_and_inventory_do_not_create_line_amount():
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Varer", None, "Antall", "Bestand"])
    sheet.append([None, "Varenr", None, None])
    sheet.append(["Produkt A", 10001, 2, 12])
    sheet.append(["Produkt B", 10002, 3, 8])
    output = BytesIO()
    workbook.save(output)

    rows = parse_xlsx_rows(output.getvalue(), "antall-og-bestand.xlsx")

    assert [row["quantity_purchased"] for row in rows] == [2.0, 3.0]
    assert [row["price_before_discount"] for row in rows] == [None, None]
    assert [row["competitor_line_amount"] for row in rows] == [None, None]
    assert {row["competitor"] for row in rows} == {""}


def test_xlsx_multiline_explicit_price_header_is_used():
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Varer", None, "Pris"])
    sheet.append([None, "Varenr", None])
    sheet.append(["Produkt A", 10001, 99.0])
    output = BytesIO()
    workbook.save(output)

    rows = parse_xlsx_rows(output.getvalue(), "eksplisitt-pris.xlsx")

    assert rows[0]["price_before_discount"] == 99.0
    assert rows[0]["calculated_unit_price"] == 99.0
    assert rows[0]["competitor"] == ""


@pytest.mark.parametrize(
    "price_headers",
    [
        ["GyMo", "GyMo"],
        ["Pris", "Pris"],
        ["Pris", "GyMo"],
    ],
)
def test_xlsx_multiline_duplicate_price_candidates_fail_closed(price_headers):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Varer", None, *price_headers])
    sheet.append([None, "Varenr", None, None])
    sheet.append(["Produkt A", 10001, 10.0, 20.0])
    output = BytesIO()
    workbook.save(output)

    rows = parse_xlsx_rows(output.getvalue(), "doble-priskolonner.xlsx")

    assert rows[0]["price_before_discount"] is None
    assert rows[0]["calculated_unit_price"] is None
    assert rows[0]["competitor_line_amount"] is None
    assert rows[0]["competitor"] == ""


def test_xlsx_multiline_explicit_price_and_source_in_same_column_are_unambiguous():
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Varer", None, "GyMo"])
    sheet.append([None, "Varenr", "Pris"])
    sheet.append(["Produkt A", 10001, 99.0])
    output = BytesIO()
    workbook.save(output)

    rows = parse_xlsx_rows(output.getvalue(), "pris-og-kilde.xlsx")

    assert rows[0]["price_before_discount"] == 99.0
    assert rows[0]["calculated_unit_price"] == 99.0
    assert rows[0]["competitor"] == "GyMo"


def test_xlsx_multiline_secondary_quantity_header_owns_price_source_column():
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Varer", None, "Antall", "GyMo"])
    sheet.append([None, "Varenr", None, "Antall"])
    sheet.append(["Produkt A", 10001, 2, 12])
    output = BytesIO()
    workbook.save(output)

    rows = parse_xlsx_rows(output.getvalue(), "sekundaert-antall.xlsx")

    assert rows[0]["quantity_purchased"] == 2.0
    assert rows[0]["price_before_discount"] is None
    assert rows[0]["competitor_line_amount"] is None
    assert rows[0]["competitor"] == ""


def test_xlsx_multiline_source_with_unknown_header_in_same_column_fails_closed():
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Varer", None, "GyMo"])
    sheet.append([None, "Varenr", "Bestand"])
    sheet.append(["Produkt A", 10001, 12])
    output = BytesIO()
    workbook.save(output)

    rows = parse_xlsx_rows(output.getvalue(), "gymo-og-bestand.xlsx")

    assert rows[0]["quantity_purchased"] == 0.0
    assert rows[0]["price_before_discount"] is None
    assert rows[0]["competitor_line_amount"] is None
    assert rows[0]["competitor"] == ""


@pytest.mark.parametrize("owned_header", ["Varenr", "Antall"])
def test_xlsx_multiline_price_source_in_owned_column_fails_closed(owned_header):
    workbook = Workbook()
    sheet = workbook.active
    if owned_header == "Varenr":
        sheet.append(["Varer", "GyMo"])
        sheet.append([None, "Varenr"])
        sheet.append(["Produkt A", 10001])
    else:
        sheet.append(["Varer", None, "GyMo"])
        sheet.append([None, "Varenr", "Antall"])
        sheet.append(["Produkt A", 10001, 2])
    output = BytesIO()
    workbook.save(output)

    rows = parse_xlsx_rows(output.getvalue(), "kolliderende-priskilde.xlsx")

    assert rows[0]["price_before_discount"] is None
    assert rows[0]["calculated_unit_price"] is None
    assert rows[0]["competitor_line_amount"] is None
    assert rows[0]["competitor"] == ""


@pytest.mark.parametrize("source_position", ["before", "between", "after"])
def test_xlsx_multiline_price_source_is_found_through_entire_header_block(source_position):
    workbook = Workbook()
    sheet = workbook.active
    if source_position == "before":
        sheet.append([None, None, "GyMo"])
        sheet.append(["Varer", None, None])
        sheet.append([None, "Varenr", None])
    elif source_position == "between":
        sheet.append(["Varer", None, None])
        sheet.append([None, None, "GyMo"])
        sheet.append([None, "Varenr", None])
    else:
        sheet.append(["Varer", None, None])
        sheet.append([None, "Varenr", None])
        sheet.append([None, None, "GyMo"])
    sheet.append(["Produkt A", 10001, 99.0])
    output = BytesIO()
    workbook.save(output)

    rows = parse_xlsx_rows(output.getvalue(), "flyttet-priskilde.xlsx")

    assert rows[0]["price_before_discount"] == 99.0
    assert rows[0]["calculated_unit_price"] == 99.0
    assert rows[0]["competitor"] == "GyMo"


def test_xlsx_multiline_late_source_is_found_before_bold_product_with_artnr():
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Varer", None, None])
    sheet.append([None, "Varenr", None])
    sheet.append([None, None, "GyMo"])
    sheet.append(["Produkt A", 10001, 99.0])
    sheet["A4"].font = Font(bold=True)
    output = BytesIO()
    workbook.save(output)

    rows = parse_xlsx_rows(output.getvalue(), "fet-produktrekke.xlsx")

    assert rows[0]["price_before_discount"] == 99.0
    assert rows[0]["competitor"] == "GyMo"


def test_xlsx_multiline_late_source_is_found_before_bold_product_without_artnr():
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Varer", None, None])
    sheet.append([None, "Varenr", None])
    sheet.append([None, None, "GyMo"])
    sheet.append(["Produkt A", None, 99.0])
    sheet["A4"].font = Font(bold=True)
    output = BytesIO()
    workbook.save(output)

    rows = parse_xlsx_rows(output.getvalue(), "fet-produkt-uten-varenummer.xlsx")

    assert len(rows) == 1
    assert rows[0]["description"] == "Produkt A"
    assert rows[0]["competitor_artnr"] == ""
    assert rows[0]["price_before_discount"] == 99.0
    assert rows[0]["competitor"] == "GyMo"


@pytest.mark.parametrize("price_headers", [["GyMo", "GyMo"], ["Pris", "Pris"]])
def test_xlsx_multiline_bold_product_survives_ambiguous_price_candidates(
    price_headers,
):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Varer", None, *price_headers])
    sheet.append([None, "Varenr", None, None])
    sheet.append(["Produkt A", None, 10.0, 20.0])
    sheet["A3"].font = Font(bold=True)
    output = BytesIO()
    workbook.save(output)

    rows = parse_xlsx_rows(output.getvalue(), "tvetydig-fet-vare.xlsx")

    assert len(rows) == 1
    assert rows[0]["description"] == "Produkt A"
    assert rows[0]["price_before_discount"] is None
    assert rows[0]["calculated_unit_price"] is None
    assert rows[0]["competitor_line_amount"] is None
    assert rows[0]["competitor"] == ""


def test_xlsx_multiline_bold_product_survives_owned_price_source_column():
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Varer", None, "GyMo"])
    sheet.append([None, "Varenr", "Antall"])
    sheet.append(["Produkt A", None, 2])
    sheet["A3"].font = Font(bold=True)
    output = BytesIO()
    workbook.save(output)

    rows = parse_xlsx_rows(output.getvalue(), "eid-priskilde-fet-vare.xlsx")

    assert len(rows) == 1
    assert rows[0]["description"] == "Produkt A"
    assert rows[0]["quantity_purchased"] == 2.0
    assert rows[0]["price_before_discount"] is None
    assert rows[0]["calculated_unit_price"] is None
    assert rows[0]["competitor_line_amount"] is None
    assert rows[0]["competitor"] == ""


@pytest.mark.parametrize("price_headers", [["GyMo", "GyMo"], ["Pris", "Pris"]])
def test_xlsx_multiline_later_bold_product_survives_ambiguous_price_candidates(
    price_headers,
):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Varer", None, *price_headers])
    sheet.append([None, "Varenr", None, None])
    sheet.append(["Produkt A", 10001, 10.0, 20.0])
    sheet.append(["Seksjon", None, None, None])
    sheet["A4"].font = Font(bold=True)
    sheet.append(["Produkt B", None, 30.0, 40.0])
    sheet["A5"].font = Font(bold=True)
    output = BytesIO()
    workbook.save(output)

    rows = parse_xlsx_rows(output.getvalue(), "sen-tvetydig-fet-vare.xlsx")

    assert [row["description"] for row in rows] == ["Produkt A", "Produkt B"]
    assert rows[1]["section"] == "Seksjon"
    assert rows[1]["price_before_discount"] is None
    assert rows[1]["calculated_unit_price"] is None
    assert rows[1]["competitor_line_amount"] is None
    assert rows[1]["competitor"] == ""


def test_xlsx_multiline_later_bold_product_survives_owned_price_source_column():
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Varer", None, "GyMo"])
    sheet.append([None, "Varenr", "Antall"])
    sheet.append(["Produkt A", 10001, 1])
    sheet.append(["Seksjon", None, None])
    sheet["A4"].font = Font(bold=True)
    sheet.append(["Produkt B", None, 2])
    sheet["A5"].font = Font(bold=True)
    output = BytesIO()
    workbook.save(output)

    rows = parse_xlsx_rows(output.getvalue(), "sen-eid-priskilde-fet-vare.xlsx")

    assert [row["description"] for row in rows] == ["Produkt A", "Produkt B"]
    assert rows[1]["section"] == "Seksjon"
    assert rows[1]["quantity_purchased"] == 2.0
    assert rows[1]["price_before_discount"] is None
    assert rows[1]["calculated_unit_price"] is None
    assert rows[1]["competitor_line_amount"] is None
    assert rows[1]["competitor"] == ""


@pytest.mark.parametrize(
    ("price_header", "expected_competitor"),
    [("GyMo", "GyMo"), ("Pris", "")],
)
def test_xlsx_multiline_price_signal_after_anchor_scan_precedes_bold_product(
    price_header,
    expected_competitor,
):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Varer", None, None])
    sheet.append([None, "Varenr", None])
    for row_number in range(3, 11):
        sheet.append([f"Seksjon {row_number}", None, None])
        sheet.cell(row_number, 1).font = Font(bold=True)
    sheet.append([None, None, price_header])
    sheet.append(["Produkt A", None, 99.0])
    sheet["A12"].font = Font(bold=True)
    output = BytesIO()
    workbook.save(output)

    rows = parse_xlsx_rows(output.getvalue(), "sent-prissignal.xlsx")

    assert len(rows) == 1
    assert rows[0]["description"] == "Produkt A"
    assert rows[0]["competitor_artnr"] == ""
    assert rows[0]["price_before_discount"] == 99.0
    assert rows[0]["competitor"] == expected_competitor


def test_xlsx_multiline_late_source_is_found_before_artnr_only_product():
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Varer", None, None])
    sheet.append([None, "Varenr", None])
    sheet.append([None, None, "GyMo"])
    sheet.append([None, 10001, 99.0])
    output = BytesIO()
    workbook.save(output)

    rows = parse_xlsx_rows(output.getvalue(), "bare-varenummer.xlsx")

    assert rows[0]["competitor_artnr"] == "10001"
    assert rows[0]["price_before_discount"] == 99.0
    assert rows[0]["competitor"] == "GyMo"


def test_xlsx_multiline_unknown_label_above_price_is_not_competitor():
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Varer", None, "NOK"])
    sheet.append([None, "Varenr", "Pris"])
    sheet.append(["Produkt A", 10001, 99.0])
    output = BytesIO()
    workbook.save(output)

    rows = parse_xlsx_rows(output.getvalue(), "valuta-og-pris.xlsx")

    assert rows[0]["price_before_discount"] == 99.0
    assert rows[0]["competitor"] == ""


@pytest.mark.parametrize("description_header", ["Varer ", "Varer\N{NO-BREAK SPACE}"])
def test_xlsx_positional_fallback_preserves_whitespace_headers(description_header):
    workbook = BytesIO()
    pd.DataFrame(
        [["Steril bandasje", None, "Ukjent verdi"]],
        columns=[description_header, "", "Ukjent kolonne"],
    ).to_excel(workbook, index=False, engine="openpyxl")

    rows = parse_xlsx_rows(workbook.getvalue(), "ukjent-format.xlsx")

    assert [row["description"] for row in rows] == ["Steril bandasje"]
